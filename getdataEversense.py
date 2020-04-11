import requests, json, arrow, hashlib, urllib, datetime
from secret import	NS_URL, NS_SECRET,	CLIENT_ID, CLIENT_SECRET,REFRESH_TOKEN
import csv
import glob
import os
import xml.etree.ElementTree
from datetime import datetime, timedelta
import calendar
import time
import argparse
import sys
import openpyxl



def get_entries_from_xlsx():
	out_treatments = []
	book = openpyxl.load_workbook('sensorglucoseresults.xlsx')

	sheet = book.active

	i=2

	while(True):
		time = sheet.cell(row=i, column=7).value
		value = sheet.cell(row=i, column=8).value
		dela= sheet.cell(row=i, column=9).value
		i=i+1
		if value == None:
			break

		start=time
		start = arrow.get(start).to("Europe/Berlin")
		date = start.format('YYYY-MM-DDTHH:mm:ssZ')
		out_treatments.append({
						"date": start.timestamp * 1000,
						"dateString": date,
						"sysTime": date,
						"device": "Eversense-Exporter",
						"rawbg": value,
						"sgv":value,
						"delta": dela,
						"filtered": value*1000,
    					"unfiltered": value*1000,
						"type":"sgv"
						})
		if len(out_treatments)>100:
			upload_nightscout_entries(out_treatments)
			out_treatments = []

	return out_treatments

def to_mgdl(mmol):
	return round(mmol*18)



def date_to_ms(ts):
	"""
	Takes a datetime object and returns POSIX UTC in milisecons
	"""
	return calendar.timegm(ts.utctimetuple()) * int(1e3)

def upload_nightscout_entries(ns_format):
	upload = requests.post(NS_URL + 'api/v1/entries?api_secret=' + NS_SECRET, json=ns_format, headers={
		'Accept': 'application/json',
		'Content-Type': 'application/json',
		'api-secret': hashlib.sha1(NS_SECRET.encode()).hexdigest()
	})
	print("Nightscout upload status:", upload.status_code, upload.text)

def upload_nightscout_treatments(ns_format):
	upload = requests.post(NS_URL + 'api/v1/treatments?api_secret=' + NS_SECRET, json=ns_format, headers={
		'Accept': 'application/json',
		'Content-Type': 'application/json',
		'api-secret': hashlib.sha1(NS_SECRET.encode()).hexdigest()
	})
	print("Nightscout upload status:", upload.status_code, upload.text)


def get_last_nightscout():
	#last = requests.get(NS_URL + 'api/v1/treatments?count=1&find[eventType]='+urllib.parse.quote('Exercise'), headers={
	last = requests.get(NS_URL + 'api/v1/treatments?count=1&find[enteredBy]='+urllib.parse.quote(NS_AUTHOR), headers={
		'Accept': 'application/json',
		'Content-Type': 'application/json',
		'api-secret': hashlib.sha1(NS_SECRET.encode()).hexdigest()
	})
	if last.status_code == 200:
		js = last.json()
		if len(js) > 0:
			duration = 0
			if 'duration' in js[0]:
				duration = js[0]['duration']
			return arrow.get(js[0]['created_at']).datetime + timedelta(minutes=duration)

def main():
	print("Getting data...", datetime.now())

	#ns_last = get_last_nightscout()

	entries = get_entries_from_xlsx()
	upload_nightscout_entries(entries)
	
	print("Got Data for days: ", len(entries))


if __name__ == '__main__':
	main()
